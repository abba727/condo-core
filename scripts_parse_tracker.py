from pathlib import Path
from datetime import datetime, date
from decimal import Decimal
from openpyxl import load_workbook
import json
import math

source = Path('/home/ubuntu/condocore_web/ProjectTracker-712Driggs.xlsx')
out = Path('/home/ubuntu/condocore_web/tracker_extracted.json')
summary_out = Path('/home/ubuntu/condocore_web/tracker_summary.txt')


def clean(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, float):
        if math.isfinite(v) and v.is_integer():
            return int(v)
        return v
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


def non_empty(row):
    return any(clean(v) is not None for v in row)

wb = load_workbook(source, data_only=True)
extracted = {"source_file": str(source), "sheets": wb.sheetnames}

# Dashboard key-value pairs
ws = wb['Dashboard']
dashboard = {}
for row in ws.iter_rows(values_only=True):
    vals = [clean(v) for v in row]
    if vals and vals[0] is not None:
        dashboard[str(vals[0])] = vals[1] if len(vals) > 1 else None
extracted['dashboard'] = dashboard

# Simple rectangular tables with row 1 headers
for sheet_name in ['Team', 'Budget', 'Expenses', 'Contracts', 'Lookup']:
    ws = wb[sheet_name]
    rows = [[clean(v) for v in row] for row in ws.iter_rows(values_only=True) if non_empty(row)]
    headers = rows[0] if rows else []
    records = []
    for row in rows[1:]:
        rec = {}
        for i, h in enumerate(headers):
            if h is None:
                continue
            rec[str(h)] = row[i] if i < len(row) else None
        if any(v is not None for v in rec.values()):
            records.append(rec)
    extracted[sheet_name.lower().replace(' ', '_')] = records

# Project Plan: metadata rows, task table, and weekly gantt marks
ws = wb['Project Plan']
project_plan_meta = {
    'project_title': clean(ws['D3'].value),
    'general_contractor': clean(ws['D4'].value),
    'project_start_date': clean(ws['D5'].value),
}
week_dates = []
for col in range(9, ws.max_column + 1):
    val = clean(ws.cell(row=6, column=col).value)
    if val is not None:
        week_dates.append({'column': col, 'date': val, 'week_number': clean(ws.cell(row=8, column=col).value)})

tasks = []
for r in range(9, ws.max_row + 1):
    wbs = clean(ws.cell(row=r, column=2).value)
    title = clean(ws.cell(row=r, column=3).value)
    if wbs is None and title is None:
        continue
    owner = clean(ws.cell(row=r, column=4).value)
    start = clean(ws.cell(row=r, column=5).value)
    due = clean(ws.cell(row=r, column=6).value)
    duration = clean(ws.cell(row=r, column=7).value)
    pct = clean(ws.cell(row=r, column=8).value)
    marks = []
    for wd in week_dates:
        mark = clean(ws.cell(row=r, column=wd['column']).value)
        if mark is not None:
            marks.append({'week_number': wd['week_number'], 'date': wd['date'], 'mark': mark})
    tasks.append({
        'row': r,
        'wbs': str(wbs) if wbs is not None else None,
        'title': title,
        'owner': owner,
        'start_date': start,
        'due_date': due,
        'duration_days': duration,
        'pct_complete': pct,
        'gantt_marks': marks,
    })
extracted['project_plan'] = {'meta': project_plan_meta, 'week_dates': week_dates, 'tasks': tasks}

# Insurances & Permits has two tables in one sheet
ws = wb['Insurances & Permits']
insurance_headers = [clean(ws.cell(row=3, column=c).value) for c in range(1, 12)]
insurance_records = []
for r in range(4, 14):
    row = [clean(ws.cell(row=r, column=c).value) for c in range(1, 12)]
    if any(v is not None for v in row):
        insurance_records.append({insurance_headers[i]: row[i] for i in range(len(insurance_headers)) if insurance_headers[i]})
permit_headers = [clean(ws.cell(row=15, column=c).value) for c in range(1, 10)]
permit_records = []
for r in range(16, ws.max_row + 1):
    row = [clean(ws.cell(row=r, column=c).value) for c in range(1, 10)]
    if any(v is not None for v in row):
        permit_records.append({permit_headers[i]: row[i] for i in range(len(permit_headers)) if permit_headers[i]})
extracted['insurances'] = insurance_records
extracted['permits'] = permit_records

out.write_text(json.dumps(extracted, indent=2), encoding='utf-8')

summary_lines = []
summary_lines.append(f"Workbook sheets: {', '.join(wb.sheetnames)}")
summary_lines.append(f"Dashboard project name in workbook: {dashboard.get('Project Name')}")
summary_lines.append(f"Dashboard address in workbook: {dashboard.get('Address')}")
summary_lines.append(f"Project Plan title in workbook: {project_plan_meta.get('project_title')}")
summary_lines.append(f"Project Plan GC: {project_plan_meta.get('general_contractor')}")
summary_lines.append(f"Project Plan start: {project_plan_meta.get('project_start_date')}")
summary_lines.append(f"Project Plan tasks extracted: {len(tasks)}")
summary_lines.append(f"Project Plan gantt weeks extracted: {len(week_dates)}")
summary_lines.append(f"Team contacts extracted: {len(extracted['team'])}")
summary_lines.append(f"Budget rows extracted: {len(extracted['budget'])}")
summary_lines.append(f"Expense rows extracted: {len(extracted['expenses'])}")
summary_lines.append(f"Contract rows extracted: {len(extracted['contracts'])}")
summary_lines.append(f"Insurance rows extracted: {len(insurance_records)}")
summary_lines.append(f"Permit rows extracted: {len(permit_records)}")
summary_lines.append('\nFirst 20 plan tasks:')
for task in tasks[:20]:
    summary_lines.append(f"- {task['wbs']} | {task['title']} | {task['start_date']} -> {task['due_date']} | pct={task['pct_complete']} | marks={len(task['gantt_marks'])}")
summary_out.write_text('\n'.join(summary_lines) + '\n', encoding='utf-8')
print(summary_out)
print(out)
