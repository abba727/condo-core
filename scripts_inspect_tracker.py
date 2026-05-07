from pathlib import Path
from openpyxl import load_workbook

source = Path('/home/ubuntu/condocore_web/ProjectTracker-712Driggs.xlsx')
wb = load_workbook(source, data_only=True, read_only=True)
print(f'Workbook: {source}')
print(f'Sheets: {wb.sheetnames}')
for ws in wb.worksheets:
    print('\n=== SHEET:', ws.title, '===')
    print('dimensions:', ws.max_row, 'rows x', ws.max_column, 'cols')
    rows = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = [v for v in row]
        if any(v is not None and str(v).strip() != '' for v in vals):
            rows.append((idx, vals))
        if len(rows) >= 12:
            break
    for idx, vals in rows:
        compact = []
        for v in vals[:20]:
            if v is None:
                compact.append('')
            else:
                s = str(v).replace('\n', ' ').strip()
                if len(s) > 80:
                    s = s[:77] + '...'
                compact.append(s)
        print(f'row {idx}:', compact)
